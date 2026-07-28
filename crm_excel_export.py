"""Export CRM analysis to Excel: daily workbook + cumulative summary with charts."""

from __future__ import annotations

import json
import os
import re
import time
import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.crm_schemas import CrmAnalysisReport

SCORE_LABELS: dict[str, str] = {
    "needs_id": "Выявление потребности",
    "objections_handled": "Отработка возражений",
    "value_presented": "Подсветка ценностей",
    "cta": "CTA",
    "deal_closed": "Закрытие сделки",
}

CHECKLIST_LABELS: dict[str, str] = {
    "greeting_contact": "Приветствие и контакт",
    "needs_identified": "Потребность выявлена",
    "context_segmentation": "Сегментация (дом/бизнес)",
    "price_in_context": "Цена в контексте",
    "values_highlighted": "Ценности озвучены",
    "objections_handled": "Возражения отработаны",
    "concrete_cta": "Конкретный CTA",
    "next_step_fixed": "Следующий шаг зафиксирован",
    "no_price_chaos": "Нет ценового хаоса",
    "response_pace_ok": "Темп ответа OK",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _save_workbook_atomic(wb: Workbook, path: Path, *, retries: int = 3) -> Path:
    """Save workbook via temp file + os.replace; clear error if Excel locks the target."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.stem}.{uuid.uuid4().hex[:8]}.tmp.xlsx")
    last_err: BaseException | None = None
    try:
        for attempt in range(1, retries + 1):
            try:
                wb.save(tmp)
                os.replace(tmp, path)
                return path
            except PermissionError as exc:
                last_err = exc
                if attempt < retries:
                    time.sleep(0.4 * attempt)
                    continue
                raise PermissionError(
                    f"Не удалось записать {path.name}: файл открыт в Excel. "
                    f"Закройте {path.name} и повторите."
                ) from last_err
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
    raise RuntimeError(f"Failed to save {path}")  # pragma: no cover


def _format_duration(seconds: float | None) -> str:
    if seconds is None:
        return "—"
    if seconds < 60:
        return f"{seconds:.0f} сек"
    if seconds < 3600:
        return f"{seconds / 60:.1f} мин"
    return f"{seconds / 3600:.1f} ч"


def _auto_width(ws, min_w: float = 10, max_w: float = 55) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                length = max(length, max(len(line) for line in lines))
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _write_title(ws, title: str, row: int = 1) -> None:
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT


def _build_recommendations(report: CrmAnalysisReport) -> str:
    agg = report.aggregate
    parts: list[str] = []
    needs = agg.avg_scores.get("needs_id", 0)
    cta = agg.avg_scores.get("cta", 0)
    if needs < 3:
        parts.append(
            f"Потребности ({needs}/5): системно не выявляются — "
            "внедрить вопрос «სახლისთვის თუ ბიზნესისთვის?» до цены."
        )
    if cta < 3:
        parts.append(
            f"CTA ({cta}/5): слабое закрытие — "
            "запретить «მოგვწერეთ ბიუჯეტი» без рекомендации."
        )
    rt = agg.response_time
    if getattr(rt, "over_sla_work", 0) and rt.over_sla_work > 0:
        parts.append(
            f"Скорость (рабочее время): {rt.over_sla_work} ответов дольше 2 мин — "
            "настроить уведомления CRM."
        )
    elif rt.over_15min > 0:
        parts.append(
            f"Скорость: {rt.over_15min} ответов дольше 15 мин — "
            "настроить уведомления в CRM."
        )
    low_checklist = [c for c, p in agg.checklist_pass_rate.items() if p < 50]
    if low_checklist:
        parts.append(f"Чеклист: слабые зоны — {', '.join(low_checklist[:5])}.")
    if not parts:
        parts.append("В целом показатели в норме; закрепить лучшие практики.")
    return "\n".join(f"• {p}" for p in parts)


def write_daily_excel(report: CrmAnalysisReport, path: Path) -> Path:
    """Daily workbook with themed tabs."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    agg = report.aggregate
    rt = agg.response_time
    date_str = report.meta.target_date

    # --- Сводка ---
    ws = wb.active
    ws.title = "Сводка"
    _write_title(ws, f"CRM-анализ ДимКава — {date_str}")
    rows = [
        ("Дата", date_str),
        ("Диалогов", report.meta.dialogs_count),
        ("Сообщений", report.meta.messages_count),
        ("LLM", f"{report.meta.llm_provider} / {report.meta.model}"),
        ("", ""),
        ("Метрика", "Значение"),
    ]
    for key, label in SCORE_LABELS.items():
        rows.append((label, agg.avg_scores.get(key, "—")))
    rows += [
        ("", ""),
        ("--- Рабочее время (10–18, пн–пт) ---", ""),
        ("Медиана (рабочие сек)", _format_duration(rt.median_work_seconds)),
        ("Среднее (рабочие сек)", _format_duration(rt.avg_work_seconds)),
        ("Ответов в рабочее время", rt.responses_count_work),
        ("Нарушений SLA >2 мин", rt.over_sla_work),
        ("", ""),
        ("--- Вне рабочего времени ---", ""),
        ("Медиана (календарное)", _format_duration(rt.median_off_seconds)),
        ("Среднее (календарное)", _format_duration(rt.avg_off_seconds)),
        ("Ответов вне смены", rt.responses_count_off),
        ("Нарушений SLA >15 мин", rt.over_sla_off),
        ("", ""),
        ("--- Общее (календарное) ---", ""),
        ("Медиана", _format_duration(rt.median_seconds)),
        ("Среднее", _format_duration(rt.avg_seconds)),
        ("Мин", f"{_format_duration(rt.min_seconds)} ({rt.min_dialog or '—'})"),
        ("Макс", f"{_format_duration(rt.max_seconds)} ({rt.max_dialog or '—'})"),
        ("Пауз >15 мин", rt.over_15min),
        ("Пауз >1 ч", rt.over_1hour),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    ws.cell(row=8, column=1).font = SUBTITLE_FONT
    ws.cell(row=8, column=2).font = SUBTITLE_FONT
    _auto_width(ws, min_w=22, max_w=45)

    # --- Оценки ---
    ws = wb.create_sheet("Оценки")
    ws.append(["Критерий", "Средний балл (0–5)", "Интерпретация"])
    _style_header_row(ws, 1, 3)
    for key, label in SCORE_LABELS.items():
        val = agg.avg_scores.get(key, 0)
        note = "Хорошо" if val >= 3 else ("Средне" if val >= 2 else "Критично")
        ws.append([label, val, note])
    _auto_width(ws)

    # --- Чеклист ---
    ws = wb.create_sheet("Чеклист")
    ws.append(["Критерий", "Прохождение %", "Статус"])
    _style_header_row(ws, 1, 3)
    for crit, label in CHECKLIST_LABELS.items():
        pct = agg.checklist_pass_rate.get(crit, 0)
        status = "OK" if pct >= 50 else "Слабо"
        ws.append([label, pct, status])
    _auto_width(ws)

    # --- Скорость ---
    ws = wb.create_sheet("Скорость")
    ws.append(
        [
            "Клиент", "Платформа", "Ответов",
            "Ср. общее", "Ср. рабоч.", "Ср. вне смены",
            "SLA>2м", "SLA>15м",
        ]
    )
    _style_header_row(ws, 1, 8)
    for d in sorted(report.dialogs, key=lambda x: x.response_time.avg_seconds or 0, reverse=True):
        rt_d = d.response_time
        ws.append([
            d.person_name,
            d.platform,
            rt_d.responses_count,
            _format_duration(rt_d.avg_seconds),
            _format_duration(rt_d.avg_work_seconds),
            _format_duration(rt_d.avg_off_seconds),
            rt_d.over_sla_work,
            rt_d.over_sla_off,
        ])
    ws.append([])
    ws.append([
        "Итого по дню", "", rt.responses_count,
        _format_duration(rt.avg_seconds),
        _format_duration(rt.avg_work_seconds),
        _format_duration(rt.avg_off_seconds),
        rt.over_sla_work,
        rt.over_sla_off,
    ])
    _auto_width(ws)

    # --- Ошибки ---
    ws = wb.create_sheet("Ошибки")
    ws.append(["#", "Ошибка", "Кол-во", "%"])
    _style_header_row(ws, 1, 4)
    for i, item in enumerate(agg.top_errors[:20], 1):
        ws.append([i, item["error"], item["count"], item["percent"]])
    _auto_width(ws, max_w=70)

    # --- Плюсы ---
    ws = wb.create_sheet("Плюсы")
    ws.append(["#", "Сильная сторона", "Кол-во", "%"])
    _style_header_row(ws, 1, 4)
    for i, item in enumerate(agg.top_strengths[:20], 1):
        ws.append([i, item["strength"], item["count"], item["percent"]])
    _auto_width(ws, max_w=70)

    # --- Диалоги ---
    ws = wb.create_sheet("Диалоги")
    headers = [
        "Клиент", "Платформа", "Сообщений",
        "Needs", "Возраж.", "Ценность", "CTA", "Сделка",
        "Суть", "Killer phrase", "Плюсы", "Ошибки",
        "Avg ответ", "Min", "Max",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for d in sorted(report.dialogs, key=lambda x: x.person_name):
        sc = d.analysis.scores
        rt_d = d.response_time
        ws.append([
            d.person_name,
            d.platform,
            d.message_count,
            sc.needs_id,
            sc.objections_handled,
            sc.value_presented,
            sc.cta,
            sc.deal_closed,
            d.analysis.summary,
            d.analysis.killer_phrase,
            "\n".join(d.analysis.strengths_found[:5]),
            "\n".join(d.analysis.errors_found[:5]),
            _format_duration(rt_d.avg_seconds),
            _format_duration(rt_d.min_seconds),
            _format_duration(rt_d.max_seconds),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
    _auto_width(ws, min_w=12, max_w=50)
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["K"].width = 40
    ws.column_dimensions["L"].width = 40

    # --- Рекомендации ---
    ws = wb.create_sheet("Рекомендации")
    rec_text = _build_recommendations(report)
    ws.cell(row=1, column=1, value="Выводы и рекомендации").font = TITLE_FONT
    ws.cell(row=3, column=1, value=rec_text)
    ws.cell(row=3, column=1).alignment = WRAP
    ws.column_dimensions["A"].width = 90
    ws.append([])
    ws.cell(row=ws.max_row + 2, column=1, value="Сценарий действий для команды").font = SUBTITLE_FONT
    actions = [
        "Утренняя планёрка (15 мин): 2–3 худших диалога по needs и CTA",
        "Правило цены: на «ფასი?» — один вопрос (дом/бизнес), потом цена",
        "Правило CTA: конкретный шаг (время, звонок, ссылка, визит)",
        "Скорость: в рабочее время (10–18) — цель ≤2 мин; вне смены — ≤15 мин",
        "Чеклист: MANAGER_PLAYBOOK_GE.md перед отправкой",
    ]
    r = ws.max_row + 1
    for i, act in enumerate(actions, 1):
        ws.cell(row=r, column=1, value=f"{i}. {act}")
        r += 1

    _save_workbook_atomic(wb, path)
    return path


def load_report_from_json(path: Path) -> CrmAnalysisReport:
    data = json.loads(path.read_text(encoding="utf-8"))
    return CrmAnalysisReport.model_validate(data)


_DAILY_REPORT_RE = re.compile(r"^crm_report_(\d{4}-\d{2}-\d{2})\.json$")


def load_all_daily_reports(output_dir: Path) -> list[CrmAnalysisReport]:
    reports: list[CrmAnalysisReport] = []
    for p in sorted(output_dir.glob("crm_report_*.json")):
        if not _DAILY_REPORT_RE.match(p.name):
            continue
        try:
            reports.append(load_report_from_json(p))
        except Exception:
            continue
    reports.sort(key=lambda r: r.meta.target_date)
    return reports


def load_daily_reports_in_range(
    output_dir: Path,
    date_from: str,
    date_to: str,
) -> list[CrmAnalysisReport]:
    reports = load_all_daily_reports(output_dir)
    return [r for r in reports if date_from <= r.meta.target_date <= date_to]


def _trend_arrow(current: float, previous: float | None, higher_is_better: bool = True) -> str:
    if previous is None:
        return "—"
    diff = current - previous
    if abs(diff) < 0.05:
        return "→"
    if higher_is_better:
        return "↑" if diff > 0 else "↓"
    return "↓" if diff > 0 else "↑"


def write_master_excel(reports: list[CrmAnalysisReport], path: Path) -> Path:
    """Cumulative workbook: daily table + trend charts."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- По дням (flat table) ---
    ws = wb.active
    ws.title = "По дням"
    headers = [
        "Дата", "Диалогов", "Сообщений",
        *SCORE_LABELS.values(),
        "Медиана общ. (сек)", "Медиана раб. (сек)", "SLA>2м", "SLA>15м вне смены",
        "Пауз >15м", "Пауз >1ч",
        "CTA чеклист %", "Needs чеклист %",
    ]
    for key in SCORE_LABELS:
        headers.append(f"Δ {SCORE_LABELS[key][:12]}")
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    prev_scores: dict[str, float] | None = None
    for report in reports:
        agg = report.aggregate
        rt = agg.response_time
        row: list[Any] = [
            report.meta.target_date,
            report.meta.dialogs_count,
            report.meta.messages_count,
        ]
        for key in SCORE_LABELS:
            row.append(agg.avg_scores.get(key, 0))
        row += [
            rt.median_seconds or 0,
            rt.median_work_seconds or 0,
            rt.over_sla_work,
            rt.over_sla_off,
            rt.over_15min,
            rt.over_1hour,
            agg.checklist_pass_rate.get("concrete_cta", 0),
            agg.checklist_pass_rate.get("needs_identified", 0),
        ]
        for key in SCORE_LABELS:
            cur = agg.avg_scores.get(key, 0)
            prev = prev_scores.get(key) if prev_scores else None
            row.append(_trend_arrow(cur, prev, higher_is_better=True))
        ws.append(row)
        prev_scores = dict(agg.avg_scores)

    _auto_width(ws)

    # --- Динамика (chart-ready numeric columns) ---
    ws_dyn = wb.create_sheet("Динамика")
    dyn_headers = ["Дата"] + list(SCORE_LABELS.values()) + [
        "Медиана общ. (мин)", "Медиана раб. (мин)", "CTA %", "Needs %", "Deal %",
    ]
    ws_dyn.append(dyn_headers)
    _style_header_row(ws_dyn, 1, len(dyn_headers))
    for report in reports:
        agg = report.aggregate
        rt = report.aggregate.response_time
        row = [report.meta.target_date]
        for key in SCORE_LABELS:
            row.append(agg.avg_scores.get(key, 0))
        row += [
            round((rt.median_seconds or 0) / 60, 1),
            round((rt.median_work_seconds or 0) / 60, 1),
            agg.checklist_pass_rate.get("concrete_cta", 0),
            agg.checklist_pass_rate.get("needs_identified", 0),
            agg.avg_scores.get("deal_closed", 0) * 20,
        ]
        ws_dyn.append(row)
    _auto_width(ws_dyn)

    # --- Графики ---
    ws_ch = wb.create_sheet("Графики")
    ws_ch.cell(row=1, column=1, value="Динамика показателей CRM").font = TITLE_FONT
    n = len(reports)
    if n >= 1:
        # Line chart: scores
        chart = LineChart()
        chart.title = "Средние оценки по дням"
        chart.y_axis.title = "Балл (0–5)"
        chart.x_axis.title = "День"
        chart.style = 10
        chart.width = 18
        chart.height = 10
        cats = Reference(ws_dyn, min_col=1, min_row=2, max_row=n + 1)
        for col_offset, key in enumerate(SCORE_LABELS.keys(), start=2):
            data = Reference(ws_dyn, min_col=col_offset, min_row=1, max_row=n + 1)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_ch.add_chart(chart, "A3")

        # Line chart: median response time (minutes) — lower is better
        chart2 = LineChart()
        chart2.title = "Медиана ответа: общее vs рабочее время (мин)"
        chart2.y_axis.title = "Минуты"
        chart2.style = 11
        chart2.width = 18
        chart2.height = 8
        med_col = 2 + len(SCORE_LABELS)
        med_work_col = med_col + 1
        data2 = Reference(ws_dyn, min_col=med_col, min_row=1, max_row=n + 1)
        data2w = Reference(ws_dyn, min_col=med_work_col, min_row=1, max_row=n + 1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.add_data(data2w, titles_from_data=True)
        chart2.set_categories(cats)
        ws_ch.add_chart(chart2, "A22")

        if n >= 1:
            # Bar chart: latest day checklist
            last = reports[-1]
            ws_ch.cell(row=40, column=1, value=f"Чеклист — {last.meta.target_date}").font = SUBTITLE_FONT
            chk_start = 41
            ws_ch.cell(row=chk_start, column=1, value="Критерий")
            ws_ch.cell(row=chk_start, column=2, value="%")
            for i, (crit, label) in enumerate(CHECKLIST_LABELS.items(), 1):
                ws_ch.cell(row=chk_start + i, column=1, value=label)
                ws_ch.cell(row=chk_start + i, column=2,
                            value=last.aggregate.checklist_pass_rate.get(crit, 0))
            bar = BarChart()
            bar.type = "bar"
            bar.title = "Чеклист последнего дня"
            bar.y_axis.title = "%"
            bar.width = 16
            bar.height = 12
            data_bar = Reference(ws_ch, min_col=2, min_row=chk_start, max_row=chk_start + len(CHECKLIST_LABELS))
            cats_bar = Reference(ws_ch, min_col=1, min_row=chk_start + 1, max_row=chk_start + len(CHECKLIST_LABELS))
            bar.add_data(data_bar, titles_from_data=True)
            bar.set_categories(cats_bar)
            ws_ch.add_chart(bar, "D40")

    # --- Легенда трендов ---
    ws_leg = wb.create_sheet("Легенда")
    ws_leg.append(["Символ", "Значение"])
    ws_leg.append(["↑", "Улучшение vs предыдущий день"])
    ws_leg.append(["↓", "Ухудшение vs предыдущий день"])
    ws_leg.append(["→", "Без изменений"])
    ws_leg.append(["—", "Первый день в серии"])
    _auto_width(ws_leg)

    _save_workbook_atomic(wb, path)
    return path


def write_period_excel(
    period_report: CrmAnalysisReport,
    daily_reports: list[CrmAnalysisReport],
    path: Path,
) -> Path:
    """Period summary workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    agg = period_report.aggregate
    rt = agg.response_time
    date_from = period_report.meta.date_from or period_report.meta.target_date
    date_to = period_report.meta.date_to or period_report.meta.target_date

    ws = wb.active
    ws.title = "Итог периода"
    _write_title(ws, f"CRM — период {date_from} … {date_to}")
    rows = [
        ("Период", f"{date_from} — {date_to}"),
        ("Дней", period_report.meta.days_in_period or len(daily_reports)),
        ("Диалогов", period_report.meta.dialogs_count),
        ("Сообщений", period_report.meta.messages_count),
        ("", ""),
    ]
    for key, label in SCORE_LABELS.items():
        rows.append((label, agg.avg_scores.get(key, "—")))
    rows += [
        ("", ""),
        ("Медиана раб. (сек)", _format_duration(rt.median_work_seconds)),
        ("SLA>2 мин (рабочее)", rt.over_sla_work),
        ("SLA>15 мин (вне смены)", rt.over_sla_off),
        ("Медиана общ. (сек)", _format_duration(rt.median_seconds)),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    _auto_width(ws, min_w=28, max_w=50)

    ws_days = wb.create_sheet("По дням")
    headers = ["Дата", "Диалогов"] + list(SCORE_LABELS.values()) + [
        "Мед. раб.", "SLA>2м", "Мед. общ.", "CTA %",
    ]
    ws_days.append(headers)
    _style_header_row(ws_days, 1, len(headers))
    for report in daily_reports:
        a = report.aggregate
        r = a.response_time
        ws_days.append([
            report.meta.target_date,
            report.meta.dialogs_count,
            *[a.avg_scores.get(k, 0) for k in SCORE_LABELS],
            _format_duration(r.median_work_seconds),
            r.over_sla_work,
            _format_duration(r.median_seconds),
            a.checklist_pass_rate.get("concrete_cta", 0),
        ])
    _auto_width(ws_days)

    ws_sp = wb.create_sheet("Скорость")
    ws_sp.append([
        "Дата", "Ответов", "Мед. раб.", "SLA>2м", "Мед. вне смены", "SLA>15м",
        "Мед. общ.", "Пауз >15м", "Пауз >1ч",
    ])
    _style_header_row(ws_sp, 1, 9)
    for report in daily_reports:
        r = report.aggregate.response_time
        ws_sp.append([
            report.meta.target_date,
            r.responses_count,
            _format_duration(r.median_work_seconds),
            r.over_sla_work,
            _format_duration(r.median_off_seconds),
            r.over_sla_off,
            _format_duration(r.median_seconds),
            r.over_15min,
            r.over_1hour,
        ])
    ws_sp.append([
        "ИТОГО", rt.responses_count,
        _format_duration(rt.median_work_seconds), rt.over_sla_work,
        _format_duration(rt.median_off_seconds), rt.over_sla_off,
        _format_duration(rt.median_seconds), rt.over_15min, rt.over_1hour,
    ])
    _auto_width(ws_sp)

    ws_err = wb.create_sheet("Ошибки недели")
    ws_err.append(["#", "Ошибка", "Кол-во", "%"])
    _style_header_row(ws_err, 1, 4)
    for i, item in enumerate(agg.top_errors[:25], 1):
        ws_err.append([i, item["error"], item["count"], item["percent"]])
    _auto_width(ws_err, max_w=70)

    ws_plus = wb.create_sheet("Плюсы недели")
    ws_plus.append(["#", "Сильная сторона", "Кол-во", "%"])
    _style_header_row(ws_plus, 1, 4)
    for i, item in enumerate(agg.top_strengths[:25], 1):
        ws_plus.append([i, item["strength"], item["count"], item["percent"]])
    _auto_width(ws_plus, max_w=70)

    if len(daily_reports) >= 1:
        ws_dyn = wb.create_sheet("Динамика")
        ws_dyn.append(["Дата"] + list(SCORE_LABELS.values()) + ["Мед. раб. (мин)", "Мед. общ. (мин)"])
        _style_header_row(ws_dyn, 1, 2 + len(SCORE_LABELS) + 2)
        for report in daily_reports:
            a = report.aggregate
            r = a.response_time
            ws_dyn.append([
                report.meta.target_date,
                *[a.avg_scores.get(k, 0) for k in SCORE_LABELS],
                round((r.median_work_seconds or 0) / 60, 1),
                round((r.median_seconds or 0) / 60, 1),
            ])
        ws_ch = wb.create_sheet("Графики")
        ws_ch.cell(row=1, column=1, value="Динамика за период").font = TITLE_FONT
        n = len(daily_reports)
        chart = LineChart()
        chart.title = "Оценки по дням"
        chart.width = 18
        chart.height = 10
        cats = Reference(ws_dyn, min_col=1, min_row=2, max_row=n + 1)
        for col_offset, key in enumerate(SCORE_LABELS.keys(), start=2):
            data = Reference(ws_dyn, min_col=col_offset, min_row=1, max_row=n + 1)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_ch.add_chart(chart, "A3")

    _save_workbook_atomic(wb, path)
    return path


def export_crm_excel(report: CrmAnalysisReport, excel_dir: Path) -> dict[str, str]:
    """Write daily + update master summary.

    Master failure after a successful daily write is not swallowed — callers see
    PermissionError with a clear «close Excel» message.
    """
    date_str = report.meta.target_date
    daily_path = excel_dir / f"CRM_DAILY_{date_str}.xlsx"
    master_path = excel_dir / "CRM_SUMMARY.xlsx"

    write_daily_excel(report, daily_path)

    by_date = {r.meta.target_date: r for r in load_all_daily_reports(excel_dir.parent)}
    by_date[report.meta.target_date] = report
    all_reports = sorted(by_date.values(), key=lambda r: r.meta.target_date)
    write_master_excel(all_reports, master_path)

    return {"daily": str(daily_path), "master": str(master_path)}


def export_excel_from_existing_json(
    output_dir: Path,
    target_date: str | None = None,
) -> dict[str, str]:
    """Rebuild Excel from saved JSON without LLM."""
    excel_dir = output_dir / "crm_excel"
    if target_date:
        json_path = output_dir / f"crm_report_{target_date}.json"
        if not json_path.exists():
            raise FileNotFoundError(json_path)
        report = load_report_from_json(json_path)
        return export_crm_excel(report, excel_dir)

    all_reports = load_all_daily_reports(output_dir)
    if not all_reports:
        raise FileNotFoundError(f"No crm_report_*.json in {output_dir}")
    paths: dict[str, str] = {}
    for report in all_reports:
        paths = export_crm_excel(report, excel_dir)
    master_path = excel_dir / "CRM_SUMMARY.xlsx"
    write_master_excel(all_reports, master_path)
    paths["master"] = str(master_path)
    return paths
